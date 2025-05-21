from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
import re
from rapidfuzz import process

from google.oauth2 import service_account
from googleapiclient.discovery import build

# ====== Config Google Sheets API ======
SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']

SA1 = r"C:\Users\Asus\OneDrive\เอกสาร\Jobber\SSL\Python Code\Search_products\search-products-459304-16109be1f71e.json"
ID1 = "124bHRo_xyV39gUytA-TXoWCbvwGPNuZ1oDtVq8gCBIY"
RANGE1 = "Sheet!A1:Z2000"  # ปรับชื่อชีทให้ถูกต้อง

SA2 = r"C:\Users\Asus\OneDrive\เอกสาร\Jobber\SSL\Python Code\Search_products\songsod02-f0faffe8b254.json"
ID2 = "1x-3hsUZ77Fx8sl_cdaFIBvdbYJDPhKQfd3VfWUnbT48"

# ====== FastAPI setup ======
app = FastAPI()

# เปิด CORS ให้เรียกจาก http://localhost:5500 ได้
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5500"],
    allow_methods=["*"],
    allow_headers=["*"],
)

def load_sheet(json_keyfile: str, spreadsheet_id: str, range_name: str) -> pd.DataFrame:
    creds = service_account.Credentials.from_service_account_file(json_keyfile, scopes=SCOPES)
    service = build('sheets', 'v4', credentials=creds)
    result = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id, range=range_name
    ).execute()
    vals = result.get('values', [])
    if not vals:
        return pd.DataFrame()
    header = vals[0]
    data = []
    for row in vals[1:]:
        if len(row) < len(header):
            row += [''] * (len(header) - len(row))
        data.append(row)
    return pd.DataFrame(data, columns=header)

def get_latest_sheet_name(json_keyfile: str, spreadsheet_id: str) -> str:
    creds = service_account.Credentials.from_service_account_file(json_keyfile, scopes=SCOPES)
    service = build('sheets', 'v4', credentials=creds)
    meta = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    sheets = meta.get('sheets', [])
    dates = []
    for s in sheets:
        title = s['properties']['title']
        if re.match(r'^\d{4}-\d{2}-\d{2}$', title):
            try:
                dates.append(pd.to_datetime(title, format='%Y-%m-%d'))
            except:
                pass
    if not dates:
        raise HTTPException(500, "ไม่พบชีทที่ชื่อเป็นวันที่")
    return max(dates).strftime('%Y-%m-%d')

# โหลดข้อมูลตลาดล่วงหน้า
df_talatthai   = load_sheet(SA1, ID1, RANGE1)
sheet2_name    = get_latest_sheet_name(SA2, ID2)
RANGE2         = f"{sheet2_name}!A1:Z2000"
df_yingcharoen = load_sheet(SA2, ID2, RANGE2)

def extract_thai(text: str) -> str:
    return ''.join(re.findall(r'[\u0E00-\u0E7F]+', str(text)))

def match_price_unit(df_market: pd.DataFrame, name: str):
    for m in df_market['ชื่อสินค้า']:
        if name in m:
            row = df_market.loc[df_market['ชื่อสินค้า'] == m].iloc[0]
            url = row.get('URL รูป', '')
            link = f'=HYPERLINK("{url}","เปิดรูป")' if url else ''
            return row['ชื่อสินค้า'], row['ราคา'], row['หน่วย'], link
    return None, None, None, None

def find_header_row(df: pd.DataFrame, possible_cols=['item', 'รายการ', 'ชื่อสินค้า']):
    for i, row in df.iterrows():
        cells = [str(c).strip().lower() for c in row.values]
        if any(col.lower() in cells for col in possible_cols):
            return i
    return None

@app.post("/upload")
async def upload_demand(file: UploadFile = File(...)):
    if not file.filename.lower().endswith((".xls", ".xlsx")):
        raise HTTPException(400, "โปรดอัปโหลดไฟล์ Excel (.xls/.xlsx)")
    content = await file.read()

    wb = load_workbook(filename=BytesIO(content), read_only=True)
    visible = [s for s in wb.sheetnames if wb[s].sheet_state == 'visible']
    wb.close()
    if not visible:
        raise HTTPException(400, "ไฟล์ไม่มีชีทที่มองเห็นได้")

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet in visible:
            df_raw = pd.read_excel(BytesIO(content), sheet_name=sheet, header=None)
            header_row = find_header_row(df_raw)
            if header_row is None:
                continue
            headers = df_raw.iloc[header_row].tolist()
            df = pd.DataFrame(df_raw.values[header_row+1:], columns=headers)

            match = process.extractOne("Item", headers, score_cutoff=80) or \
                    process.extractOne("รายการ", headers, score_cutoff=80) or \
                    process.extractOne("ชื่อสินค้า", headers, score_cutoff=80)
            if not match:
                continue
            item_col = match[0]

            df['ชื่อสินค้า_clean'] = df[item_col].apply(extract_thai)
            df[['ชื่อตลาดไท','ราคา_ตลาดไท','หน่วย_ตลาดไท','รูป_ตลาดไท']] = \
                df['ชื่อสินค้า_clean'].apply(lambda x: pd.Series(match_price_unit(df_talatthai, x)))
            df[['ชื่อยิ่งเจริญ','ราคา_ยิ่งเจริญ','หน่วย_ยิ่งเจริญ','รูป_ยิ่งเจริญ']] = \
                df['ชื่อสินค้า_clean'].apply(lambda x: pd.Series(match_price_unit(df_yingcharoen, x)))

            df.to_excel(writer, sheet_name=sheet[:31], index=False)

    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=result.xlsx"}
    )
